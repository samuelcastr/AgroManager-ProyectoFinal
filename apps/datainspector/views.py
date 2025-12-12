from django.shortcuts import render
from rest_framework import viewsets
from rest_framework.permissions import IsAuthenticated
from rest_framework.filters import SearchFilter, OrderingFilter
from django_filters.rest_framework import DjangoFilterBackend
from .models import ExportRequest
from .serializers import ExportRequestSerializer
import openpyxl
from openpyxl.chart import BarChart, Reference
from django.http import HttpResponse
from django.apps import apps
from django.views import View
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from rest_framework.views import APIView
from rest_framework.response import Response
import datetime


@method_decorator(login_required, name='dispatch')
class ExportExcelSelectView(View):
    template_name = 'datainspector/export_select.html'

    def get(self, request):
        # Listar apps disponibles
        app_labels = ['cultivos', 'inventario', 'sensores', 'datainspector']
        return render(request, self.template_name, {'apps': app_labels})

    def post(self, request):
        app_label = request.POST.get('app')
        if not app_label:
            return self.get(request)
        # Obtener modelos de la app seleccionada
        models = apps.get_app_config(app_label).get_models()
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for model in models:
            sheet = wb.create_sheet(title=model._meta.object_name)
            fields = [f.name for f in model._meta.fields]
            sheet.append(fields)
            for obj in model.objects.all():
                sheet.append([getattr(obj, f) for f in fields])
            # Gráfica simple (si hay datos numéricos)
            num_cols = [i+1 for i, f in enumerate(fields) if hasattr(model, f) and isinstance(getattr(model, f), (int, float))]
            if num_cols:
                chart = BarChart()
                data = Reference(sheet, min_col=num_cols[0], min_row=1, max_col=num_cols[-1], max_row=sheet.max_row)
                chart.add_data(data, titles_from_data=True)
                sheet.add_chart(chart, f'E5')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{app_label}_export.xlsx"'
        wb.save(response)
        return response

class ExportExcelSelectViewSet(viewsets.ModelViewSet):
    queryset = ExportRequest.objects.all().order_by('-created_at')
    serializer_class = ExportRequestSerializer
    permission_classes = [IsAuthenticated]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        app_label = serializer.validated_data['app']
        # Guardar la solicitud
        self.perform_create(serializer)
        # Generar el Excel
        models = apps.get_app_config(app_label).get_models()
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for model in models:
            sheet = wb.create_sheet(title=model._meta.object_name)
            fields = [f.name for f in model._meta.fields]
            sheet.append(fields)
            for obj in model.objects.all():
                row = []
                for f in fields:
                    value = getattr(obj, f, None)
                    # Relaciones: mostrar como string legible
                    if hasattr(value, '__str__') and not isinstance(value, (str, int, float, dict, list, type(None), datetime.date, datetime.datetime)):
                        value = str(value)
                    # Diccionario o lista
                    if isinstance(value, dict):
                        value = str(value)
                    elif isinstance(value, list):
                        value = ', '.join(map(str, value))
                    # Datetime con zona horaria
                    if isinstance(value, datetime.datetime) and value.tzinfo is not None:
                        value = value.replace(tzinfo=None)
                    # Formatear fechas y datetimes como YYYY-MM-DD
                    if isinstance(value, (datetime.datetime, datetime.date)):
                        value = value.strftime('%Y-%m-%d')
                    row.append(value)
                sheet.append(row)
            # Gráfica simple (si hay datos numéricos)
            num_cols = [i+1 for i, f in enumerate(fields) if hasattr(model, f) and isinstance(getattr(model, f, None), (int, float))]
            if num_cols:
                chart = BarChart()
                data = Reference(sheet, min_col=num_cols[0], min_row=1, max_col=num_cols[-1], max_row=sheet.max_row)
                chart.add_data(data, titles_from_data=True)
                sheet.add_chart(chart, f'E5')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{app_label}_export.xlsx"'
        wb.save(response)
        return response

class ExportGlobalExcelViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    def list(self, request):
        app_labels = ['cultivos', 'inventario', 'sensores', 'datainspector']
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for app_label in app_labels:
            models = apps.get_app_config(app_label).get_models()
            for model in models:
                sheet = wb.create_sheet(title=f'{app_label}_{model._meta.object_name}')
                fields = [f.name for f in model._meta.fields]
                sheet.append(fields)
                for obj in model.objects.all():
                    row = []
                    for f in fields:
                        value = getattr(obj, f, None)
                        # Relaciones: mostrar como string legible
                        if hasattr(value, '__str__') and not isinstance(value, (str, int, float, dict, list, type(None), datetime.date, datetime.datetime)):
                            value = str(value)
                        # Diccionario o lista
                        if isinstance(value, dict):
                            value = str(value)
                        elif isinstance(value, list):
                            value = ', '.join(map(str, value))
                        # Datetime con zona horaria
                        if isinstance(value, datetime.datetime) and value.tzinfo is not None:
                            value = value.replace(tzinfo=None)
                        # Formatear fechas y datetimes como YYYY-MM-DD
                        if isinstance(value, (datetime.datetime, datetime.date)):
                            value = value.strftime('%Y-%m-%d')
                        row.append(value)
                    sheet.append(row)
                # Gráfica simple (si hay datos numéricos)
                num_cols = [i+1 for i, f in enumerate(fields) if hasattr(model, f) and isinstance(getattr(model, f, None), (int, float))]
                if num_cols:
                    chart = BarChart()
                    data = Reference(sheet, min_col=num_cols[0], min_row=1, max_col=num_cols[-1], max_row=sheet.max_row)
                    chart.add_data(data, titles_from_data=True)
                    sheet.add_chart(chart, f'E5')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="global_export.xlsx"'
        wb.save(response)
        return response
